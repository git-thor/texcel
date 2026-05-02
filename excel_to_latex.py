#!/usr/bin/env python3
"""
Excel to LaTeX Table Converter

Converts Excel spreadsheets to LaTeX table format with support for:
- Math notation preservation ($\\gamma=0$, $\\pi_H(s)$, etc.)
- Bold/italic formatting from Excel
- Multiline output format for readability
- Automatic column header comments
- Smart escaping of special LaTeX characters
- Empty row/column handling (midrule and separators)

Example:
    >>> excel_to_latex("table.xlsx", "table.tex")
    [OK] LaTeX table written to: table.tex
"""

import pandas as pd
import argparse
import re
from pathlib import Path
from typing import Optional, Union, Set, Tuple, List
from openpyxl import load_workbook


def excel_to_latex(
    excel_path: str,
    output_path: str,
    sheet_name: Union[int, str] = 0,
    use_tabularx: bool = True,
    landscape: bool = True,
    first_col_width: str = "2cm",
    font_size: Optional[str] = None,
    caption: Optional[str] = None,
    label: Optional[str] = None,
    skip_rows: int = 0,
    escape_special_chars: bool = True,
    preserve_formatting: bool = True,
    multiline: bool = True,
) -> str:
    """
    Convert Excel file to LaTeX table, preserving cell formatting.

    Args:
        excel_path: Path to input Excel file (.xlsx)
        output_path: Path to output .tex file
        sheet_name: Excel sheet name or index (default: 0)
        use_tabularx: Use tabularx (auto-width) vs standard tabular (default: True)
        landscape: Wrap in landscape environment (default: True)
        first_col_width: Width of first column for tabularx (default: "2cm")
        font_size: LaTeX font size (tiny, scriptsize, footnotesize, small, normalsize)
        caption: Table caption for \\caption{}
        label: LaTeX label for cross-referencing (e.g., "tab:mytable")
        skip_rows: Number of rows to skip at beginning (default: 0)
        escape_special_chars: Escape &, %, $, #, _, {, }, ~, ^, \\ (default: True)
        preserve_formatting: Preserve bold/italic formatting from Excel (default: True)
        multiline: Use multiline format with column comments (default: True)

    Returns:
        str: Generated LaTeX table content

    Raises:
        FileNotFoundError: If excel_path does not exist
        KeyError: If sheet_name not found in workbook

    Example:
        >>> latex = excel_to_latex("data.xlsx", "output.tex", caption="My Table")
        [OK] LaTeX table written to: output.tex
    """
    # Load workbook to access formatting
    wb = load_workbook(excel_path, data_only=True)

    # Get sheet
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    # Read all cells with formatting
    data: List[List[str]] = []
    bold_rows: Set[int] = set()
    bold_cells: Set[Tuple[int, int]] = set()
    col_has_content: dict = {}  # Track which columns have content

    for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
        if row_idx < skip_rows:
            continue

        row_data: List[str] = []
        row_has_bold = False
        row_is_empty = True

        for col_idx, cell in enumerate(row):
            value = cell.value if cell.value is not None else ""
            row_data.append(str(value))

            if value:  # Check if cell has content
                row_is_empty = False
                col_has_content[col_idx] = True

            # Check if cell is bold
            if preserve_formatting and cell.font and cell.font.bold:
                row_has_bold = True
                bold_cells.add((row_idx - skip_rows, col_idx))

        # Mark empty rows for midrule generation
        if row_is_empty and row_data:
            data.append(("__EMPTY_ROW__",))  # Special marker for midrule
        elif row_data:
            data.append(row_data)
            if row_has_bold:
                bold_rows.add(row_idx - skip_rows)

    wb.close()

    # Create DataFrame
    df = pd.DataFrame(data)

    # Determine which columns are fully empty (separator columns)
    # Only middle columns (not leading/trailing) become separators
    num_cols = df.shape[1]
    empty_columns: Set[int] = set()
    for col_idx in range(num_cols):
        if col_idx not in col_has_content:
            # Check if this is a middle column (has data columns on both sides)
            has_left_data = any(
                i < col_idx and i in col_has_content for i in range(num_cols)
            )
            has_right_data = any(
                i > col_idx and i in col_has_content for i in range(num_cols)
            )
            if has_left_data and has_right_data:
                empty_columns.add(col_idx)

    # Generate table content
    latex_content = generate_latex_table(
        df,
        bold_rows=bold_rows,
        bold_cells=bold_cells,
        empty_columns=empty_columns,
        use_tabularx=use_tabularx,
        landscape=landscape,
        first_col_width=first_col_width,
        font_size=font_size,
        caption=caption,
        label=label,
        escape_special_chars=escape_special_chars,
        multiline=multiline,
    )

    # Write to file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(latex_content)

    print(f"[OK] LaTeX table written to: {output_path}")
    print(f"  Dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
    print(f"  Bold rows: {len(bold_rows)}, Bold cells: {len(bold_cells)}")

    return latex_content


def generate_latex_table(
    df: pd.DataFrame,
    bold_rows: Optional[Set[int]] = None,
    bold_cells: Optional[Set[Tuple[int, int]]] = None,
    empty_columns: Optional[Set[int]] = None,
    use_tabularx: bool = True,
    landscape: bool = True,
    first_col_width: str = "2cm",
    font_size: Optional[str] = None,
    caption: Optional[str] = None,
    label: Optional[str] = None,
    escape_special_chars: bool = True,
    multiline: bool = True,
) -> str:
    """
    Generate LaTeX table code from DataFrame.

    Args:
        df: pandas DataFrame containing table data
        bold_rows: Set of row indices that should be bold
        bold_cells: Set of (row, col) tuples for bold cells
        empty_columns: Set of column indices that are empty (separators)
        use_tabularx: Use tabularx environment (default: True)
        landscape: Wrap in landscape environment (default: True)
        first_col_width: Width of first column for tabularx (default: "2cm")
        font_size: LaTeX font size command
        caption: Table caption
        label: LaTeX label
        escape_special_chars: Escape special LaTeX characters (default: True)
        multiline: Use multiline format with comments (default: True)

    Returns:
        str: Complete LaTeX table code including environment wrappers

    Example:
        >>> latex = generate_latex_table(df, caption="Results", multiline=True)
    """
    if empty_columns is None:
        empty_columns = set()

    num_cols = df.shape[1]

    # Build column specification with | for empty columns
    if use_tabularx:
        col_parts: List[str] = []
        for col_idx in range(num_cols):
            if col_idx in empty_columns:
                col_parts.append("|")  # Separator column
            else:
                if col_idx == 0:
                    col_parts.append(
                        f">{{\\raggedright\\arraybackslash}}p{{{first_col_width}}}"
                    )
                else:
                    col_parts.append(">{{\\centering\\arraybackslash}}X")
        col_spec = "@{}" + "".join(col_parts) + "@{{}}"
    else:
        # All columns centered, with | for empty columns
        col_spec = "@{}"
        for col_idx in range(num_cols):
            if col_idx in empty_columns:
                col_spec += "|"
            else:
                col_spec += "c"
        col_spec += "@{}"

    latex: List[str] = []

    # Landscape wrapper
    if landscape:
        latex.extend(
            [
                "\\clearpage",
                "\\begin{landscape}",
            ]
        )

    # Table environment
    latex.extend(
        [
            "\\begin{table}[p]",
            "    \\centering",
        ]
    )

    # Font size
    if font_size:
        latex.append(f"    \\{font_size}")

    # Caption and label
    if caption:
        latex.append(f"    \\caption{{{caption}}}")
    if label:
        latex.append(f"    \\label{{{label}}}")

    # Array stretch
    latex.append("    \\renewcommand{\\arraystretch}{1.2}")

    # Begin tabular
    table_env = "tabularx" if use_tabularx else "tabular"
    width_spec = "\\linewidth" if use_tabularx else ""
    latex.append(f"    \\begin{{{table_env}}}{{{width_spec}}}{{{col_spec}}}")
    latex.append("        \\toprule")

    # Convert DataFrame to LaTeX rows
    df_filled = df.fillna("")

    # Extract header row for column comments (first row)
    header_row: Optional[List[str]] = None
    if len(df_filled) > 0:
        header_row = df_filled.iloc[0].tolist()

    for idx, row in df_filled.iterrows():
        row_values: List[str] = row.tolist()

        # Check if this is an empty row marker (first cell is the marker)
        if row_values and row_values[0] == "__EMPTY_ROW__":
            latex.append("        \\midrule")
            continue

        # Escape special LaTeX characters if enabled
        if escape_special_chars:
            # Preserve math mode $...$ while escaping other characters
            row_escaped = [escape_latex(val, preserve_math=True) for val in row_values]
        else:
            # Even with escape_special_chars=False, we must escape % to prevent LaTeX comments
            # But preserve math notation
            row_escaped: List[str] = []
            for val in row_values:
                val_str = str(val)
                # Always escape % (LaTeX comment character)
                # Use regex to avoid escaping \% that's already escaped or % inside $...$

                # First, extract math parts
                math_parts: List[str] = []

                def save_math(match: re.Match) -> str:
                    math_parts.append(match.group(0))
                    return f"MATHPLACEHOLDER{len(math_parts) - 1}ENDMATH"

                val_temp = re.sub(r"(?<!\\)\$([^$]+?)(?<!\\)\$", save_math, val_str)

                # Escape % in non-math parts
                val_escaped = val_temp.replace("%", "\\%")

                # Restore math parts
                for i, math_part in enumerate(math_parts):
                    val_escaped = val_escaped.replace(
                        f"MATHPLACEHOLDER{i}ENDMATH", math_part
                    )

                row_escaped.append(val_escaped)

        # Apply bold formatting
        if bold_cells:
            row_formatted: List[str] = []
            for col_idx, val in enumerate(row_escaped):
                if (idx, col_idx) in bold_cells:
                    row_formatted.append(f"\\textbf{{{val}}}")
                else:
                    row_formatted.append(val)
            row_escaped = row_formatted

        # Format row - include empty columns as empty cells
        if multiline:
            # Each column on a new line for readability
            if row_escaped:
                # First cell with single indent
                latex.append("            " + row_escaped[0])
                # Remaining cells with & prefix, value, and column header comment
                for col_idx in range(1, len(row_escaped)):
                    val = row_escaped[col_idx]
                    # Get column header for comment
                    col_comment = ""
                    if header_row and col_idx < len(header_row):
                        header_val = str(header_row[col_idx]).strip()
                        if header_val:
                            col_comment = f"  % {header_val}"
                    latex.append("            &    " + val + col_comment)
            latex.append("            \\\\")
            # Add blank line after each row for better readability
            latex.append("")
        else:
            row_str = "        " + " & ".join(row_escaped) + " \\\\"
            latex.append(row_str)

    latex.append("        \\bottomrule")
    latex.append(f"    \\end{{{table_env}}}")
    latex.append("\\end{table}")

    # End landscape
    if landscape:
        latex.extend(
            [
                "\\end{landscape}",
                "\\clearpage",
            ]
        )

    return "\n".join(latex)


def escape_latex(text: Union[str, None], preserve_math: bool = True) -> str:
    """
    Escape special LaTeX characters.

    Args:
        text: Text to escape
        preserve_math: If True, preserve math mode $...$ without escaping
            the dollar signs (default: True)

    Returns:
        str: Escaped text safe for LaTeX

    Example:
        >>> escape_latex("100% complete")
        '100\\%'
        >>> escape_latex("$\\gamma=0$")
        '$\\gamma=0$'
        >>> escape_latex("$\\gamma=0$", preserve_math=False)
        '\\$\\gamma=0\\$'
    """
    if text is None or text == "":
        return ""

    text = str(text)

    # Special marker for empty rows - don't escape it
    if text == "__EMPTY_ROW__":
        return text

    # If preserving math, extract math parts first
    math_parts: List[str] = []
    if preserve_math and "$" in text:
        # Find all $...$ math expressions

        def save_math(match: re.Match) -> str:
            math_parts.append(match.group(0))
            # Use placeholder that won't be affected by escaping
            return f"MATHPLACEHOLDER{len(math_parts) - 1}ENDMATH"

        # Match $...$ but not \$
        text = re.sub(r"(?<!\\)\$([^$]+?)(?<!\\)\$", save_math, text)

    # Escape backslash FIRST (order matters!)
    text = text.replace("\\", "\\textbackslash{}")

    # Then escape other special characters (but NOT $ if preserving math)
    replacements: List[Tuple[str, str]] = [
        ("&", "\\&"),
        ("%", "\\%"),
        ("#", "\\#"),
        ("_", "\\_"),
        ("{", "\\{"),
        ("}", "\\}"),
        ("~", "\\textasciitilde{}"),
        ("^", "\\textasciicircum{}"),
    ]

    # Only escape $ if NOT preserving math
    if not preserve_math:
        replacements.insert(2, ("$", "\\$"))

    for old, new in replacements:
        text = text.replace(old, new)

    # Restore math parts
    for i, math_part in enumerate(math_parts):
        text = text.replace(f"MATHPLACEHOLDER{i}ENDMATH", math_part)

    return text


def list_sheets(excel_path: str) -> None:
    """
    List all sheets in Excel file with dimensions.

    Args:
        excel_path: Path to Excel file

    Example:
        >>> list_sheets("data.xlsx")

        Sheets in 'data.xlsx':
          [0] Sheet1: 10 rows × 5 columns
          [1] Sheet2: 20 rows × 3 columns
    """
    wb = load_workbook(excel_path, data_only=True)
    print(f"\nSheets in '{excel_path}':")
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = len(list(ws.iter_rows(values_only=True)))
        cols = ws.max_column
        print(f"  [{i}] {sheet_name}: {rows} rows × {cols} columns")
    wb.close()


def main() -> None:
    """Main entry point for command-line interface."""
    parser = argparse.ArgumentParser(
        description="Convert Excel data to LaTeX tables (preserves formatting)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # List sheets in Excel file
  python excel_to_latex.py data.xlsx --list-sheets

  # Basic conversion (tabularx, landscape, preserves bold)
  python excel_to_latex.py data.xlsx -o table.tex

  # With caption and label
  python excel_to_latex.py data.xlsx -o table.tex -c "My Table" -l "tab:mytable"

  # Standard tabular (not tabularx), portrait
  python excel_to_latex.py data.xlsx -o table.tex --no-tabularx --no-landscape

  # Custom first column width, font size
  python excel_to_latex.py data.xlsx -o table.tex -w "3cm" -f footnotesize

  # Skip intro rows, don't escape special chars
  python excel_to_latex.py data.xlsx -o table.tex --skip-rows 2 --no-escape
        """,
    )

    parser.add_argument("excel_path", help="Path to input Excel file (.xlsx)")
    parser.add_argument(
        "-o", "--output", default="output_table.tex", help="Output .tex file path"
    )
    parser.add_argument(
        "-s", "--sheet", default=0, help="Sheet name or index (default: 0)"
    )
    parser.add_argument(
        "--list-sheets", action="store_true", help="List all sheets and exit"
    )
    parser.add_argument(
        "-t",
        "--tabularx",
        action="store_true",
        default=True,
        help="Use tabularx (default)",
    )
    parser.add_argument(
        "--no-tabularx",
        action="store_false",
        dest="tabularx",
        help="Use standard tabular",
    )
    parser.add_argument(
        "-l",
        "--landscape",
        action="store_true",
        default=True,
        help="Landscape mode (default)",
    )
    parser.add_argument(
        "--no-landscape", action="store_false", dest="landscape", help="Portrait mode"
    )
    parser.add_argument(
        "-w", "--width", default="2cm", help="First column width (default: 2cm)"
    )
    parser.add_argument(
        "-f",
        "--font",
        choices=["tiny", "scriptsize", "footnotesize", "small", "normalsize"],
        help="Font size",
    )
    parser.add_argument("-c", "--caption", default="", help="Table caption")
    parser.add_argument("--label", default="", help="LaTeX label (e.g., tab:mytable)")
    parser.add_argument(
        "--skip-rows", type=int, default=0, help="Skip N rows at beginning"
    )
    parser.add_argument(
        "--no-escape", action="store_true", help="Don't escape special LaTeX chars"
    )
    parser.add_argument(
        "--no-formatting",
        action="store_true",
        help="Don't preserve Excel formatting (bold, etc.)",
    )
    parser.add_argument(
        "--no-multiline",
        action="store_true",
        help="Use compact single-line format (default is multiline for readability)",
    )

    args = parser.parse_args()

    # Check file exists
    if not Path(args.excel_path).exists():
        print(f"✗ Error: File not found: {args.excel_path}")
        return

    # List sheets if requested
    if args.list_sheets:
        list_sheets(args.excel_path)
        return

    # Convert
    excel_to_latex(
        excel_path=args.excel_path,
        output_path=args.output,
        sheet_name=args.sheet,
        use_tabularx=args.tabularx,
        landscape=args.landscape,
        first_col_width=args.width,
        font_size=args.font,
        caption=args.caption if args.caption else None,
        label=args.label if args.label else None,
        skip_rows=args.skip_rows,
        escape_special_chars=not args.no_escape,
        preserve_formatting=not args.no_formatting,
        multiline=not args.no_multiline,  # Multiline is now default
    )


if __name__ == "__main__":
    main()
