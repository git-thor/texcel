#!/usr/bin/env python3
"""
LaTeX to Excel Table Converter

Converts LaTeX tabular/tabularx tables to Excel format with support for:
- Bold formatting preservation (\\textbf{})
- Math notation preservation ($\\gamma=0$, etc.)
- Comment stripping (% comments removed)
- Multiple table extraction
- Empty cell handling

Example:
    >>> latex_to_excel("table.tex", "output.xlsx")
    [OK] Excel file written to: output.xlsx
"""

import re
import argparse
from pathlib import Path
from typing import List, Tuple, Set
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def latex_to_excel(
    latex_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    skip_empty_rows: bool = False,
    preserve_formatting: bool = True,
    table_index: int = 0,
) -> None:
    """
    Convert LaTeX table to Excel file.

    Args:
        latex_path: Path to input .tex file
        output_path: Path to output Excel file (.xlsx)
        sheet_name: Name of Excel sheet (default: "Sheet1")
        skip_empty_rows: Skip rows that are empty or contain only whitespace
            (default: False - preserves empty rows for structure)
        preserve_formatting: Apply bold formatting for \\textbf{} content
            (default: True)
        table_index: Which table to extract if multiple tables exist
            (0-indexed, default: 0)

    Raises:
        FileNotFoundError: If latex_path does not exist
        IndexError: If table_index out of range

    Example:
        >>> latex_to_excel("table.tex", "output.xlsx", table_index=1)
        [OK] Excel file written to: output.xlsx
    """
    # Read LaTeX file (handle BOM and different encodings)
    try:
        with open(latex_path, "r", encoding="utf-8-sig") as f:
            latex_content = f.read()
    except UnicodeDecodeError:
        # Fallback to UTF-16 if UTF-8 fails (common on Windows)
        with open(latex_path, "r", encoding="utf-16") as f:
            latex_content = f.read()

    # Extract table data
    tables = extract_tables(latex_content)

    if not tables:
        print("✗ Error: No LaTeX tables found in file")
        return

    if table_index >= len(tables):
        print(
            f"✗ Error: Table index {table_index} out of range (found {len(tables)} tables)"
        )
        return

    table_data, bold_cells = tables[table_index]

    # Note: We don't skip empty rows here anymore - they're handled by parse_tabular_content
    # if skip_empty_rows:
    #     table_data = [row for row in table_data if any(cell.strip() for cell in row)]

    # Create Excel workbook
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = sheet_name[:31]  # Excel limit: 31 chars

    # Write data with formatting
    for row_idx, row in enumerate(table_data, start=1):
        for col_idx, cell in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell)

            # Apply bold formatting if cell was bold in LaTeX
            if preserve_formatting and (row_idx - 1, col_idx - 1) in bold_cells:
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    # Auto-adjust column widths (basic)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = min(50, max(max_length, len(str(cell.value))))
        ws.column_dimensions[column].width = max_length + 2

    # Save workbook
    wb.save(output_path)
    wb.close()

    print(f"[OK] Excel file written to: {output_path}")
    print(
        f"  Dimensions: {len(table_data)} rows x {len(table_data[0]) if table_data else 0} columns"
    )
    print(f"  Bold cells: {len(bold_cells)}")


def extract_tables(
    latex_content: str,
) -> List[Tuple[List[List[str]], Set[Tuple[int, int]]]]:
    """
    Extract all tabular/tabularx environments from LaTeX content.

    Args:
        latex_content: LaTeX document or snippet content

    Returns:
        List of (data, bold_cells) tuples, where:
            - data: 2D list of cell contents (List[List[str]])
            - bold_cells: Set of (row_idx, col_idx) tuples for bold cells

    Example:
        >>> tables = extract_tables("\\\\begin{tabular}...\\\\end{tabular}")
        >>> len(tables)
        1
    """
    tables: List[Tuple[List[List[str]], Set[Tuple[int, int]]]] = []

    # Match tabular or tabularx environments (non-greedy, handles whitespace)
    pattern = r"\\begin\{(tabular|tabularx)\}.*?(.*?)\\end\{\1\}"

    for match in re.finditer(pattern, latex_content, re.DOTALL):
        # table_env = match.group(1)
        table_content = match.group(2)
        data, bold_cells = parse_tabular_content(table_content)
        if data:
            tables.append((data, bold_cells))

    return tables


def parse_tabular_content(
    content: str,
) -> Tuple[List[List[str]], Set[Tuple[int, int]]]:
    """
    Parse tabular content into 2D array.

    Args:
        content: Content between \\begin{tabular} and \\end{tabular}
            (excluding environment markers)

    Returns:
        Tuple containing:
            - data: 2D list of cell contents
            - bold_cells: Set of (row_idx, col_idx) tuples for bold cells

    Example:
        >>> data, bold = parse_tabular_content("{cc} A & B \\\\\\\\ C & D")
        >>> data
        [['A', 'B'], ['C', 'D']]
    """
    data: List[List[str]] = []
    bold_cells: Set[Tuple[int, int]] = set()

    # Remove column specification (everything between \\begin{tabular} and first \\\\ or content)
    # Column spec always starts with { and ends with }
    match = re.match(r"^\s*\{.*?\}\s*(?=\n|$)", content, re.DOTALL)
    if match:
        content = content[match.end() :]

    # Remove \\toprule, \\midrule, \\bottomrule commands (with optional whitespace/newlines after)
    content = re.sub(r"\\(?:top|mid|bottom)rule\s*\n?", "", content)

    # Strip comments from the entire content BEFORE splitting into rows
    # This prevents comments from interfering with row parsing
    content = strip_latex_comments_full(content)

    # Split by row-ending \\\\ (must be followed by newline, whitespace, or end)
    # This avoids splitting on \\\\ inside commands
    rows = re.split(r"\\\\(?=\s|$)", content)

    for row_idx, row in enumerate(rows):
        row = row.strip()
        if not row:
            continue

        # Skip if row is a standalone command (not \\textbf, \\textit, etc.)
        if row.startswith("\\") and not row.startswith("\\text"):
            continue

        # Split by & (cell separator), but not \\& (escaped ampersand)
        cells = re.split(r"(?<!\\)&", row)
        row_data: List[str] = []

        for col_idx, cell in enumerate(cells):
            cell = cell.strip()

            # Skip if cell is a standalone command (not \\textbf, \\textit, etc.)
            if cell.startswith("\\") and not cell.startswith("\\text"):
                row_data.append("")
                continue

            # Check for bold formatting BEFORE unescaping
            is_bold = "\\textbf{" in cell

            # Remove \\textbf{} wrapper but keep content
            cell = unescape_latex(cell)

            row_data.append(cell)
            if is_bold and cell:
                bold_cells.add((row_idx, len(row_data) - 1))

        # Only skip if the row has NO cells at all (not even empty ones)
        if len(row_data) > 0:
            data.append(row_data)

    return data, bold_cells


def strip_latex_comments(text: str) -> str:
    """
    Remove LaTeX comments (everything after unescaped %).

    Handles \\% (escaped percent) correctly. For single-line use
    (strips from % to end of string).

    Args:
        text: Single line of LaTeX content

    Returns:
        str: Text with comments removed

    Example:
        >>> strip_latex_comments("Test % ignore this")
        'Test '
        >>> strip_latex_comments("100\\\\% complete")
        '100\\\\%'
    """
    if not text:
        return ""

    result: List[str] = []
    i = 0
    while i < len(text):
        # Check for escaped percent \\%
        if i < len(text) - 1 and text[i] == "\\" and text[i + 1] == "%":
            result.append("\\%")
            i += 2
        # Check for comment start
        elif text[i] == "%":
            # Everything after % is a comment, stop here
            break
        else:
            result.append(text[i])
            i += 1

    return "".join(result)


def strip_latex_comments_full(text: str) -> str:
    """
    Remove all LaTeX comments from multi-line text.

    Handles \\% (escaped percent) correctly. Processes line by line.

    Args:
        text: Multi-line LaTeX content

    Returns:
        str: Text with all comments removed

    Example:
        >>> strip_latex_comments_full("Line 1 % comment\\\\nLine 2")
        'Line 1 \\nLine 2'
    """
    if not text:
        return ""

    lines = text.split("\n")
    result_lines: List[str] = []

    for line in lines:
        result_line: List[str] = []
        i = 0
        while i < len(line):
            # Check for escaped percent \\%
            if i < len(line) - 1 and line[i] == "\\" and line[i + 1] == "%":
                result_line.append("\\%")
                i += 2
            # Check for comment start
            elif line[i] == "%":
                # Everything after % is a comment, skip rest of line
                break
            else:
                result_line.append(line[i])
                i += 1
        result_lines.append("".join(result_line))

    return "\n".join(result_lines)


def unescape_latex(text: str) -> str:
    """
    Convert LaTeX escaped characters and commands back to plain text.

    Handles \\textbf{}, math mode, and special characters.
    PRESERVES math notation (does NOT convert to Unicode).

    Args:
        text: LaTeX-formatted text

    Returns:
        str: Plain text with LaTeX commands removed, math preserved

    Example:
        >>> unescape_latex("\\\\textbf{Bold}")
        'Bold'
        >>> unescape_latex("100\\\\%")
        '100%'
        >>> unescape_latex("$\\\\gamma=0$")
        '$\\\\gamma=0$'
    """
    if not text:
        return ""

    # Remove \\textbf{} wrapper (keep content)
    text = re.sub(r"\\textbf\{([^}]*)\}", r"\1", text)

    # Remove \\textit{} wrapper (keep content)
    text = re.sub(r"\\textit\{([^}]*)\}", r"\1", text)

    # Remove \\texttt{} wrapper (keep content)
    text = re.sub(r"\\texttt\{([^}]*)\}", r"\1", text)

    # Math mode $...$ is PRESERVED (not converted to Unicode)
    # Just ensure it's properly handled

    # Unescape special characters
    replacements: List[Tuple[str, str]] = [
        ("\\&", "&"),
        ("\\%", "%"),
        ("\\#", "#"),
        ("\\_", "_"),
        ("\\{", "{"),
        ("\\}", "}"),
        ("\\textasciitilde{}", "~"),
        ("\\textasciitilde\\{\\}", "~"),
        ("\\textasciicircum{}", "^"),
        ("\\textasciicircum\\{\\}", "^"),
        ("\\textbackslash{}", "\\"),
        ("\\textbackslash\\{\\}", "\\"),
        ("\\-", "-"),  # Soft hyphen
        ("\\,", " "),  # Thin space
        ("\\;", " "),  # Medium space
        ("\\:", " "),  # Medium space
        ("\\>", " "),  # Medium space
        ("\\quad", "  "),  # Quad space
        ("\\qquad", "    "),  # Double quad
    ]

    for old, new in replacements:
        text = text.replace(old, new)

    # Convert Unicode Greek letters to LaTeX commands BEFORE removing other commands
    greek_to_latex: List[Tuple[str, str]] = [
        ("α", "\\alpha"),
        ("β", "\\beta"),
        ("γ", "\\gamma"),
        ("δ", "\\delta"),
        ("ε", "\\epsilon"),
        ("ζ", "\\zeta"),
        ("η", "\\eta"),
        ("θ", "\\theta"),
        ("ι", "\\iota"),
        ("κ", "\\kappa"),
        ("λ", "\\lambda"),
        ("μ", "\\mu"),
        ("ν", "\\nu"),
        ("ξ", "\\xi"),
        ("π", "\\pi"),
        ("ρ", "\\rho"),
        ("σ", "\\sigma"),
        ("ς", "\\varsigma"),
        ("τ", "\\tau"),
        ("υ", "\\upsilon"),
        ("φ", "\\phi"),
        ("χ", "\\chi"),
        ("ψ", "\\psi"),
        ("ω", "\\omega"),
        ("Α", "\\Alpha"),
        ("Β", "\\Beta"),
        ("Γ", "\\Gamma"),
        ("Δ", "\\Delta"),
        ("Ε", "\\Epsilon"),
        ("Ζ", "\\Zeta"),
        ("Η", "\\Eta"),
        ("Θ", "\\Theta"),
        ("Ι", "\\Iota"),
        ("Κ", "\\Kappa"),
        ("Λ", "\\Lambda"),
        ("Μ", "\\Mu"),
        ("Ν", "\\Nu"),
        ("Ξ", "\\Xi"),
        ("Π", "\\Pi"),
        ("Ρ", "\\Rho"),
        ("Σ", "\\Sigma"),
        ("Τ", "\\Tau"),
        ("Υ", "\\Upsilon"),
        ("Φ", "\\Phi"),
        ("Χ", "\\Chi"),
        ("Ψ", "\\Psi"),
        ("Ω", "\\Omega"),
    ]

    for old, new in greek_to_latex:
        text = text.replace(old, new)

    # Remove remaining LaTeX commands (like \\raggedright, \\centering, etc.)
    # BUT preserve math commands inside $...$ and Greek letter commands
    # First, extract math content
    math_parts: List[str] = []
    math_pattern = r"\$([^\$]+)\$"

    def save_math(match: re.Match) -> str:
        math_parts.append(match.group(0))
        return f"__MATH_{len(math_parts) - 1}__"

    text = re.sub(math_pattern, save_math, text)

    # Extract Greek letter LaTeX commands to preserve them
    greek_placeholders: List[str] = []

    def save_greek(match: re.Match) -> str:
        greek_placeholders.append(match.group(0))
        return f"__GREEK_{len(greek_placeholders) - 1}__"

    greek_pattern = r"\\(alpha|beta|gamma|delta|epsilon|zeta|eta|theta|iota|kappa|lambda|mu|nu|xi|pi|rho|sigma|varsigma|tau|upsilon|phi|chi|psi|omega|Alpha|Beta|Gamma|Delta|Epsilon|Zeta|Eta|Theta|Iota|Kappa|Lambda|Mu|Nu|Xi|Pi|Rho|Sigma|Tau|Upsilon|Phi|Chi|Psi|Omega)"
    text = re.sub(greek_pattern, save_greek, text)

    # Remove other LaTeX commands
    text = re.sub(r"\\[a-zA-Z]+", "", text)

    # Restore Greek commands
    for i, placeholder in enumerate(greek_placeholders):
        text = text.replace(f"__GREEK_{i}__", placeholder)

    # Restore math parts
    for i, math_part in enumerate(math_parts):
        text = text.replace(f"__MATH_{i}__", math_part)

    # Remove extra whitespace
    text = " ".join(text.split())

    return text


def list_tables(latex_path: str) -> None:
    """
    List all tables in LaTeX file with dimensions.

    Args:
        latex_path: Path to .tex file

    Example:
        >>> list_tables("tables.tex")

        Tables in 'tables.tex':
          [0] 10 rows × 5 columns, 3 bold cells
          [1] 20 rows × 3 columns, 0 bold cells
    """
    with open(latex_path, "r", encoding="utf-8") as f:
        content = f.read()

    tables = extract_tables(content)

    print(f"\nTables in '{latex_path}':")
    for i, (data, bold_cells) in enumerate(tables):
        rows = len(data)
        cols = len(data[0]) if data else 0
        print(f"  [{i}] {rows} rows × {cols} columns, {len(bold_cells)} bold cells")


def main() -> None:
    """Main entry point for command-line interface."""
    parser = argparse.ArgumentParser(
        description="Convert LaTeX tables to Excel (preserves bold formatting)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # List tables in LaTeX file
  python latex_to_excel.py table.tex --list-tables

  # Basic conversion (first table)
  python latex_to_excel.py table.tex -o output.xlsx

  # Convert specific table by index
  python latex_to_excel.py table.tex -o output.xlsx -t 1

  # Custom sheet name, no formatting
  python latex_to_excel.py table.tex -o output.xlsx -s "Data" --no-formatting

  # Keep empty rows
  python latex_to_excel.py table.tex -o output.xlsx --keep-empty
        """,
    )

    parser.add_argument("latex_path", help="Path to input .tex file")
    parser.add_argument(
        "-o", "--output", default="output.xlsx", help="Output Excel file path (.xlsx)"
    )
    parser.add_argument(
        "-s", "--sheet", default="Sheet1", help="Sheet name (default: Sheet1)"
    )
    parser.add_argument(
        "--list-tables", action="store_true", help="List all tables and exit"
    )
    parser.add_argument(
        "-t",
        "--table-index",
        type=int,
        default=0,
        help="Table index to extract (0-indexed, default: 0)",
    )
    parser.add_argument(
        "--no-formatting", action="store_true", help="Don't preserve bold formatting"
    )
    parser.add_argument("--keep-empty", action="store_true", help="Keep empty rows")

    args = parser.parse_args()

    # Check file exists
    if not Path(args.latex_path).exists():
        print(f"✗ Error: File not found: {args.latex_path}")
        return

    # List tables if requested
    if args.list_tables:
        list_tables(args.latex_path)
        return

    # Convert
    latex_to_excel(
        latex_path=args.latex_path,
        output_path=args.output,
        sheet_name=args.sheet,
        skip_empty_rows=not args.keep_empty,
        preserve_formatting=not args.no_formatting,
        table_index=args.table_index,
    )


if __name__ == "__main__":
    main()
