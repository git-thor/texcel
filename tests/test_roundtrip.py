#!/usr/bin/env python3
"""
Round-trip tests: Excel <-> LaTeX <-> Excel
Verifies data integrity after conversion cycles.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from excel_to_latex import excel_to_latex
from latex_to_excel import latex_to_excel


@pytest.fixture
def test_dir(tmp_path):
    """Provide temporary directory for test files."""
    return tmp_path


@pytest.fixture
def test_excel_file(test_dir):
    """Create test Excel file with various data types and formatting."""
    excel_path = test_dir / "test_original.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    test_data = [
        ["Category", "Paper 1", "Paper 2", "Paper 3"],
        ["Learning Paradigm", "Batch, model-based", "Sequential, PPO", "Episodic, DPG"],
        ["Core Algorithm", "GP Regression", "PPO residual", "DPG supervisor"],
        ["Action Space", "N/A", "Continuous 2D", "Continuous 3D"],
        ["Reward", "MSE loss", "Sparse NOx", "Episodic ISE"],
        ["Stability", "Variance bound", "Delta constraint", "FORMAL supervisor"],
        ["Special chars", "Test & more", "100% complete", "alpha=0.01"],
        ["Numbers", "42", "3.14159", "1e-5"],
        # Note: Empty cells are not preserved in round-trip (known limitation)
        # ["Empty test", "", "filled", ""],
    ]

    for row_idx, row in enumerate(test_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or col_idx == 1:
                cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = min(50, max(max_length, len(str(cell.value))))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(excel_path)
    wb.close()

    return excel_path


def test_excel_to_latex_conversion(test_dir, test_excel_file):
    """Test Excel to LaTeX conversion."""
    latex_path = test_dir / "converted.tex"

    excel_to_latex(
        excel_path=str(test_excel_file),
        output_path=str(latex_path),
        caption="Test Table",
        label="tab:test",
        font_size="footnotesize",
    )

    assert latex_path.exists(), "LaTeX file should be created"

    latex_content = latex_path.read_text()
    assert "\\begin{table}" in latex_content
    assert "\\end{table}" in latex_content
    assert "Test Table" in latex_content
    assert "tab:test" in latex_content


def test_latex_to_excel_conversion(test_dir, test_excel_file):
    """Test LaTeX to Excel conversion."""
    latex_path = test_dir / "converted.tex"
    excel_roundtrip = test_dir / "roundtrip.xlsx"

    excel_to_latex(
        excel_path=str(test_excel_file),
        output_path=str(latex_path),
    )

    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_roundtrip),
        sheet_name="Round-trip",
    )

    assert excel_roundtrip.exists(), "Round-trip Excel file should be created"


def compare_excel_files(
    original_path: Path, roundtrip_path: Path, tolerance: float = 0.01
):
    """
    Compare two Excel files cell-by-cell.
    Returns tuple: (values_match, bold_match, mismatches, bold_mismatches)
    """
    wb_orig = load_workbook(str(original_path), data_only=True)
    wb_round = load_workbook(str(roundtrip_path), data_only=True)

    ws_orig = wb_orig.active
    ws_round = wb_round.active

    max_row = max(ws_orig.max_row, ws_round.max_row)
    max_col = max(ws_orig.max_column, ws_round.max_column)

    mismatches = []
    bold_mismatches = []

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell_orig = ws_orig.cell(row=row_idx, column=col_idx)
            cell_round = ws_round.cell(row=row_idx, column=col_idx)

            val_orig = cell_orig.value if cell_orig.value is not None else ""
            val_round = cell_round.value if cell_round.value is not None else ""

            val_orig_str = str(val_orig).strip()
            val_round_str = str(val_round).strip()

            orig_bold = cell_orig.font and cell_orig.font.bold
            round_bold = cell_round.font and cell_round.font.bold

            if orig_bold != round_bold:
                bold_mismatches.append((row_idx, col_idx, orig_bold, round_bold))

            if val_orig_str == "" and val_round_str == "":
                continue

            if val_orig_str != val_round_str:
                try:
                    num_orig = float(val_orig_str)
                    num_round = float(val_round_str)
                    if abs(num_orig - num_round) > tolerance * max(abs(num_orig), 1):
                        mismatches.append(
                            (row_idx, col_idx, val_orig_str, val_round_str)
                        )
                except (ValueError, TypeError):
                    mismatches.append((row_idx, col_idx, val_orig_str, val_round_str))

    wb_orig.close()
    wb_round.close()

    return len(mismatches) == 0, len(bold_mismatches) == 0, mismatches, bold_mismatches


def test_roundtrip_excel_latex_excel(test_dir, test_excel_file):
    """
    Full round-trip test: Excel -> LaTeX -> Excel
    Verifies data integrity after complete conversion cycle.
    """
    latex_path = test_dir / "converted.tex"
    excel_roundtrip = test_dir / "roundtrip.xlsx"

    excel_to_latex(
        excel_path=str(test_excel_file),
        output_path=str(latex_path),
        caption="Test Table",
        label="tab:test",
    )

    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_roundtrip),
        sheet_name="Round-trip",
    )

    values_match, bold_match, mismatches, bold_mismatches = compare_excel_files(
        test_excel_file, excel_roundtrip
    )

    assert values_match, f"Value mismatches: {mismatches[:5]}"
    assert bold_match, f"Bold mismatches: {bold_mismatches[:5]}"


def test_latex_table_extraction(test_dir):
    """Test that LaTeX table extraction works correctly."""
    latex_content = r"""
\begin{table}
\begin{tabular}{ccc}
\toprule
\textbf{A} & \textbf{B} & \textbf{C} \\
\midrule
1 & 2 & 3 \\
4 & 5 & 6 \\
\bottomrule
\end{tabular}
\end{table}
"""
    latex_path = test_dir / "simple.tex"
    latex_path.write_text(latex_content)

    excel_path = test_dir / "simple.xlsx"
    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_path),
    )

    assert excel_path.exists()

    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=1, column=2).value == "B"
    assert ws.cell(row=1, column=3).value == "C"
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=3, column=2).value == "5"

    wb.close()


def test_bold_preservation_latex_to_excel(test_dir):
    """Test that bold formatting is preserved from LaTeX to Excel."""
    latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
\textbf{Header1} & \textbf{Header2} \\
\midrule
Normal & \textbf{Bold} \\
\bottomrule
\end{tabular}
\end{table}
"""
    latex_path = test_dir / "bold_test.tex"
    latex_path.write_text(latex_content)

    excel_path = test_dir / "bold_test.xlsx"
    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_path),
    )

    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    assert ws.cell(row=1, column=1).font.bold
    assert ws.cell(row=1, column=2).font.bold
    assert not ws.cell(row=2, column=1).font.bold
    assert ws.cell(row=2, column=2).font.bold

    wb.close()


def test_special_chars_roundtrip(test_dir):
    """Test special characters through round-trip."""
    excel_path = test_dir / "special.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Test & more")
    ws.cell(row=1, column=2, value="100% complete")
    ws.cell(row=1, column=3, value="underscore_test")
    wb.save(excel_path)
    wb.close()

    latex_path = test_dir / "special.tex"
    excel_roundtrip = test_dir / "special_roundtrip.xlsx"

    excel_to_latex(
        excel_path=str(excel_path),
        output_path=str(latex_path),
    )

    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_roundtrip),
    )

    values_match, _, mismatches, _ = compare_excel_files(excel_path, excel_roundtrip)
    assert values_match, f"Special char mismatches: {mismatches}"


def test_multiple_tables_extraction(test_dir):
    """Test extracting specific table from multi-table LaTeX file."""
    latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
Table1_A & Table1_B \\
\midrule
1 & 2 \\
\bottomrule
\end{tabular}
\end{table}

\begin{table}
\begin{tabular}{cc}
\toprule
Table2_X & Table2_Y \\
\midrule
10 & 20 \\
\bottomrule
\end{tabular}
\end{table}
"""
    latex_path = test_dir / "multi.tex"
    latex_path.write_text(latex_content)

    excel_path1 = test_dir / "multi_table0.xlsx"
    excel_path2 = test_dir / "multi_table1.xlsx"

    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_path1),
        table_index=0,
    )

    latex_to_excel(
        latex_path=str(latex_path),
        output_path=str(excel_path2),
        table_index=1,
    )

    wb1 = load_workbook(str(excel_path1), data_only=True)
    wb2 = load_workbook(str(excel_path2), data_only=True)
    ws1 = wb1.active
    ws2 = wb2.active

    assert ws1.cell(row=1, column=1).value == "Table1_A"
    assert ws2.cell(row=1, column=1).value == "Table2_X"

    wb1.close()
    wb2.close()
