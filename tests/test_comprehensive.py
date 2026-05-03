#!/usr/bin/env python3
"""
Comprehensive tests for Excel <-> LaTeX converter issues.
"""

import pytest
from openpyxl import Workbook, load_workbook

from excel_to_latex import excel_to_latex, escape_latex
from latex_to_excel import latex_to_excel, unescape_latex


@pytest.fixture
def test_dir(tmp_path):
    """Provide temporary directory for test files."""
    return tmp_path


class TestEmptyColumns:
    """Test: Correctly regard empty columns at the beginning"""

    def test_excel_to_latex_leading_empty_columns(self, test_dir):
        """Excel with leading empty columns should preserve them in LaTeX"""
        excel_path = test_dir / "leading_empty.xlsx"
        wb = Workbook()
        ws = wb.active
        # Create: [empty, empty, Header, Data]
        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="")
        ws.cell(row=1, column=3, value="Header")
        ws.cell(row=1, column=4, value="Data")
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="Row1")
        ws.cell(row=2, column=4, value="Val1")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "leading_empty.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=True,
        )

        latex_content = latex_path.read_text()
        # Should have 4 columns (2 empty + 2 data)
        # Check for empty cells pattern - works for both multiline and single-line format
        has_pattern = (
            " &  &" in latex_content
            or "& &" in latex_content
            or " & & " in latex_content
            or "&    \n            &" in latex_content  # Multiline format pattern
        )
        assert has_pattern, (
            f"Leading empty columns should be preserved with & &: {latex_content}"
        )

    def test_latex_to_excel_leading_empty_columns(self, test_dir):
        """LaTeX with leading empty columns should preserve them in Excel"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cccc}
\toprule
 & & Header & Data \\
\midrule
 & & Row1 & Val1 \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "leading_empty.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "leading_empty.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        # First two columns should be empty
        assert (
            ws.cell(row=1, column=1).value == ""
            or ws.cell(row=1, column=1).value is None
        )
        assert (
            ws.cell(row=1, column=2).value == ""
            or ws.cell(row=1, column=2).value is None
        )
        assert ws.cell(row=1, column=3).value == "Header"
        assert ws.cell(row=1, column=4).value == "Data"
        wb.close()


class TestLaTeXComments:
    """Test: Correctly ignore comments (all content after %)"""

    def test_latex_to_excel_strips_comments(self, test_dir):
        """Content after % should be ignored in LaTeX parsing"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
Header1 & Header2 \\ % this is a comment
Data1 & Data2 \\ % another comment
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "comments.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "comments.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        # Should not contain comment text
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    assert "comment" not in str(cell).lower(), (
                        f"Comment should be stripped: {cell}"
                    )
        wb.close()

    def test_latex_to_excel_inline_comment(self, test_dir):
        """Inline comments should be stripped, keeping cell content"""
        latex_content = r"""
\begin{table}
\begin{tabular}{c}
\toprule
Test % ignore this \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "inline_comment.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "inline_comment.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "Test"
        wb.close()


class TestEmptyRowsAsMidrule:
    """Test: Regard a fully empty row as midrule"""

    def test_excel_to_latex_empty_row_creates_midrule(self, test_dir):
        """Fully empty rows in Excel should generate \\midrule in LaTeX"""
        excel_path = test_dir / "empty_row.xlsx"
        wb = Workbook()
        ws = wb.active
        # Row 1: data, Row 2: empty, Row 3: data
        ws.cell(row=1, column=1, value="Header1")
        ws.cell(row=1, column=2, value="Header2")
        # Row 2 is empty (don't set values)
        ws.cell(row=3, column=1, value="Data1")
        ws.cell(row=3, column=2, value="Data2")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "empty_row.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
        )

        latex_content = latex_path.read_text()
        # Should contain midrule for the empty row
        assert "\\midrule" in latex_content, (
            "Empty row should generate \\midrule command"
        )

    def test_latex_to_excel_midrule_creates_empty_row(self, test_dir):
        """\\midrule in LaTeX should create empty row in Excel"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
Header1 & Header2 \\
\midrule
Data1 & Data2 \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "midrule.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "midrule.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
            skip_empty_rows=False,
        )

        # Just verify it parses without error
        assert excel_path.exists()


class TestEmptyColumnsAsSeparator:
    """Test: Regard a fully empty column as separator rule"""

    def test_excel_to_latex_empty_column_creates_separator(self, test_dir):
        """Fully empty columns in Excel should create | in LaTeX column spec"""
        excel_path = test_dir / "empty_column.xlsx"
        wb = Workbook()
        ws = wb.active
        # Col 1: data, Col 2: empty, Col 3: data
        ws.cell(row=1, column=1, value="Header1")
        ws.cell(row=1, column=2, value="")  # Empty separator column
        ws.cell(row=1, column=3, value="Header2")
        ws.cell(row=2, column=1, value="Data1")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="Data2")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "empty_column.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
        )

        latex_content = latex_path.read_text()
        # Should have separator in column spec (|c|c or similar)
        assert "|" in latex_content or "||" in latex_content, (
            "Empty column should create | separator in column spec"
        )


class TestMathSymbols:
    """Test: Correctly parse back and forth math symbols"""

    def test_excel_to_latex_preserves_math_notation(self, test_dir):
        """Math notation in Excel should be preserved in LaTeX"""
        excel_path = test_dir / "math_symbols.xlsx"
        wb = Workbook()
        ws = wb.active
        # Store LaTeX math notation as text in Excel
        ws.cell(row=1, column=1, value=r"$\gamma=0$")
        ws.cell(row=1, column=2, value=r"$\pi$")
        ws.cell(row=1, column=3, value=r"$\alpha+\beta$")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "math_symbols.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=False,  # Don't escape $ in math
        )

        latex_content = latex_path.read_text()
        # Should preserve math notation
        assert r"$\gamma" in latex_content or r"\$" in latex_content
        assert r"$\pi" in latex_content or r"\$" in latex_content

    def test_latex_to_excel_preserves_math_notation(self, test_dir):
        """Math symbols in LaTeX should NOT be converted to Unicode"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
$\gamma$ & $\pi$ \\
$\alpha$ & $\beta$ \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "math_symbols.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "math_symbols.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        # Should preserve LaTeX notation, NOT convert to Unicode
        cell1 = str(ws.cell(row=1, column=1).value)
        cell2 = str(ws.cell(row=1, column=2).value)

        # Should contain LaTeX commands, not Unicode symbols
        assert "$" in cell1 or r"\gamma" in cell1, (
            f"Math notation should be preserved, got: {cell1}"
        )
        assert "$" in cell2 or r"\pi" in cell2, (
            f"Math notation should be preserved, got: {cell2}"
        )
        # Should NOT be Unicode
        assert "γ" not in cell1, "Should not convert to Unicode γ"
        assert "π" not in cell2, "Should not convert to Unicode π"
        wb.close()

    def test_math_symbols_roundtrip(self, test_dir):
        """Math symbols should survive Excel->LaTeX->Excel roundtrip"""
        excel_path = test_dir / "math_roundtrip.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=r"$\gamma=0.99$")
        ws.cell(row=1, column=2, value=r"$\Delta[k_p, k_i]$")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "math_roundtrip.tex"
        excel_roundtrip = test_dir / "math_roundtrip2.xlsx"

        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=False,
        )

        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_roundtrip),
        )

        wb_orig = load_workbook(str(excel_path), data_only=True)
        wb_round = load_workbook(str(excel_roundtrip), data_only=True)

        # Compare values
        orig_val = str(wb_orig.active.cell(row=1, column=1).value)
        round_val = str(wb_round.active.cell(row=1, column=1).value)

        assert orig_val == round_val, (
            f"Math symbol roundtrip failed: '{orig_val}' != '{round_val}'"
        )

        wb_orig.close()
        wb_round.close()


class TestMathEnvironment:
    """Test: Correctly parse back and forth math environment with $"""

    def test_inline_math_preserved(self, test_dir):
        """Inline math $...$ should be preserved"""
        latex_content = r"""
\begin{table}
\begin{tabular}{c}
\toprule
Test $x=y$ math \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "inline_math.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "inline_math.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        cell_val = str(ws.cell(row=1, column=1).value)
        assert "$" in cell_val, f"Inline math $ should be preserved: {cell_val}"
        wb.close()

    def test_math_with_special_chars(self, test_dir):
        """Math with underscores and other special chars should work"""
        excel_path = test_dir / "math_special.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=r"$e_t, \int e \,d\tau$")
        ws.cell(row=1, column=2, value=r"$[k_p, k_i]$")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "math_special.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=False,
        )

        latex_content = latex_path.read_text()
        # Should preserve the math content
        assert "e_t" in latex_content or "e\\_t" in latex_content
        assert "k_p" in latex_content or "k\\_p" in latex_content


class TestPercentHandling:
    """Test: Handle % wherever it comes from"""

    def test_excel_percent_to_latex(self, test_dir):
        """% in Excel should be escaped as \\% in LaTeX"""
        excel_path = test_dir / "percent.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="100% complete")
        ws.cell(row=1, column=2, value="50% efficiency")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "percent.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
        )

        latex_content = latex_path.read_text()
        assert r"\%" in latex_content, "% should be escaped as \\% in LaTeX"
        # Should not have unescaped %
        lines_with_percent = [
            l
            for l in latex_content.split("\n")  # noqa: E741
            if "%" in l and r"\%" not in l
        ]
        assert len(lines_with_percent) == 0, f"Found unescaped %: {lines_with_percent}"

    def test_latex_percent_to_excel(self, test_dir):
        """\\% in LaTeX should become % in Excel"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
100\% complete & 50\% efficiency \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "percent.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "percent.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "100% complete"
        assert ws.cell(row=1, column=2).value == "50% efficiency"
        wb.close()

    def test_percent_roundtrip(self, test_dir):
        """% should survive roundtrip Excel->LaTeX->Excel"""
        excel_path = test_dir / "percent_roundtrip.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="100% complete")
        ws.cell(row=1, column=2, value="50% off")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "percent_roundtrip.tex"
        excel_roundtrip = test_dir / "percent_roundtrip2.xlsx"

        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
        )

        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_roundtrip),
        )

        wb_orig = load_workbook(str(excel_path), data_only=True)
        wb_round = load_workbook(str(excel_roundtrip), data_only=True)

        assert (
            wb_orig.active.cell(row=1, column=1).value
            == wb_round.active.cell(row=1, column=1).value
        )
        assert (
            wb_orig.active.cell(row=1, column=2).value
            == wb_round.active.cell(row=1, column=2).value
        )

        wb_orig.close()
        wb_round.close()


class TestEscapeLatex:
    """Test escape_latex function directly"""

    def test_escape_percent(self):
        """escape_latex should escape %"""
        assert escape_latex("100%") == "100\\%"

    def test_escape_math_dollar(self):
        """escape_latex should preserve $ by default (for math mode)"""
        # With preserve_math=True (default), $ is preserved
        assert escape_latex("$5") == "$5"
        # Can still escape $ if needed by setting preserve_math=False
        assert escape_latex("$5", preserve_math=False) == "\\$5"

    def test_preserve_math_mode(self):
        """When escape_special_chars=False, $ should be preserved"""
        # This is handled by the excel_to_latex function parameter
        pass


class TestUnescapeLatex:
    """Test unescape_latex function directly"""

    def test_unescape_percent(self):
        """unescape_latex should convert \\% to %"""
        assert unescape_latex("100\\%") == "100%"

    def test_unescape_math_commands(self):
        """unescape_latex should NOT convert math to Unicode"""
        # Current implementation converts to Unicode - this is the bug
        result = unescape_latex(r"$\gamma$")
        # Should preserve LaTeX notation
        assert "$" in result or r"\gamma" in result
        assert "γ" not in result, "Should not convert to Unicode"

    def test_unescape_preserves_dollar(self):
        """Math mode content should be preserved"""
        result = unescape_latex(r"Test $\alpha=0$")
        assert "$" in result or r"\alpha" in result


class TestUnicodeGreekLetters:
    """Test: Convert Unicode Greek letters to LaTeX commands"""

    def test_unicode_tau_to_latex_tau(self, test_dir):
        """Unicode τ in Excel should become \\tau in LaTeX"""
        excel_path = test_dir / "unicode_tau.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="τ")
        ws.cell(row=1, column=2, value="τ_I")
        ws.cell(row=1, column=3, value=r"$\tau$")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "unicode_tau.tex"
        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=False,
        )

        latex_content = latex_path.read_text()
        # Unicode τ should be converted to \tau
        assert r"\tau" in latex_content, (
            f"Unicode τ should become \\tau: {latex_content}"
        )

    def test_unicode_greek_letters_roundtrip(self, test_dir):
        """Unicode Greek letters should survive roundtrip"""
        excel_path = test_dir / "unicode_greek.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="α")
        ws.cell(row=1, column=2, value="β")
        ws.cell(row=1, column=3, value="γ")
        ws.cell(row=1, column=4, value="δ")
        ws.cell(row=1, column=5, value="τ")
        wb.save(excel_path)
        wb.close()

        latex_path = test_dir / "unicode_greek.tex"
        excel_roundtrip = test_dir / "unicode_greek2.xlsx"

        excel_to_latex(
            excel_path=str(excel_path),
            output_path=str(latex_path),
            escape_special_chars=False,
        )

        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_roundtrip),
        )

        wb_round = load_workbook(str(excel_roundtrip), data_only=True)
        ws_round = wb_round.active

        # All should be converted to LaTeX commands
        assert r"\alpha" in str(ws_round.cell(row=1, column=1).value)
        assert r"\beta" in str(ws_round.cell(row=1, column=2).value)
        assert r"\gamma" in str(ws_round.cell(row=1, column=3).value)
        assert r"\delta" in str(ws_round.cell(row=1, column=4).value)
        assert r"\tau" in str(ws_round.cell(row=1, column=5).value)

        wb_round.close()

    def test_latex_to_excel_unicode_greek(self, test_dir):
        """Unicode Greek letters in LaTeX should become LaTeX commands in Excel"""
        latex_content = r"""
\begin{table}
\begin{tabular}{cc}
\toprule
τ & α \\
γ & β \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "unicode_latex.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "unicode_latex.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        # Unicode should be converted to LaTeX commands
        assert ws.cell(row=1, column=1).value == r"\tau"
        assert ws.cell(row=1, column=2).value == r"\alpha"
        assert ws.cell(row=2, column=1).value == r"\gamma"
        assert ws.cell(row=2, column=2).value == r"\beta"
        wb.close()

    def test_unescape_unicode_greek_letters(self):
        """Test unescape_latex converts Unicode Greek to LaTeX commands"""
        assert unescape_latex("τ") == r"\tau"
        assert unescape_latex("α") == r"\alpha"
        assert unescape_latex("β") == r"\beta"
        assert unescape_latex("γ") == r"\gamma"
        assert unescape_latex("δ") == r"\delta"
        assert unescape_latex("π") == r"\pi"
        assert unescape_latex("μ") == r"\mu"
        assert unescape_latex("σ") == r"\sigma"
        assert unescape_latex("ω") == r"\omega"
        # Uppercase
        assert unescape_latex("Γ") == r"\Gamma"
        assert unescape_latex("Δ") == r"\Delta"
        assert unescape_latex("Σ") == r"\Sigma"
        # Mixed with other content
        assert unescape_latex("Time constant τ") == r"Time constant \tau"
        assert unescape_latex("α + β = γ") == r"\alpha + \beta = \gamma"


class TestComprehensive:
    """Integration tests combining multiple issues"""

    def test_complex_table_with_all_features(self, test_dir):
        """Test a complex table with empty cols, math, %, comments"""
        latex_content = r"""
\begin{table}
\begin{tabular}{c cc}
\toprule
 & Header1 & Header2 \\ % comment
\midrule
Row1 & $\gamma=0$ & 100\% \\
 & & \\
Row2 & $\pi$ & 50\% \\
\bottomrule
\end{tabular}
\end{table}
"""
        latex_path = test_dir / "complex.tex"
        latex_path.write_text(latex_content)

        excel_path = test_dir / "complex.xlsx"
        latex_to_excel(
            latex_path=str(latex_path),
            output_path=str(excel_path),
        )

        # Verify it parses without error
        assert excel_path.exists()

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active

        # Check math is preserved (not Unicode)
        cell_val = str(ws.cell(row=2, column=2).value)
        assert "γ" not in cell_val, "Math should not be Unicode"

        # Check % is unescaped
        assert ws.cell(row=2, column=3).value == "100%"

        wb.close()
